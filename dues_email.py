import openpyxl, smtplib, sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

unpaid_members = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

smtp_obj = smtplib.SMTP('boo@example.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
smtp_obj.login('boo@example.com', 'MY_SECRET_PASSWORD')


for name,email in unpaid_members:
    body ="Subject: %s dues unpaid.\nDear %s,\nRecords show that you've not made payments for %s.Please make this payment as soon as possible.\nThank you!" % (latest_month, name, latest_month)
    print('Sending email to %s...' % (email))
    sendmail_status = smtp_obj.sendmail('boo@example.com', email, body)
    if sendmail_status != {}:
        print('There was a problem sending mail to %s' %(email, sendmail_status))

smtp_obj.quit()

